import csv
import io
import logging
from datetime import datetime
from typing import Any

from sqlalchemy import insert, or_
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.gy import GyMagnet

logger = logging.getLogger(__name__)


def _read_rows(file_bytes: bytes) -> list[tuple]:
    """读取 Excel/CSV 行，兼容任意格式，解析失败自动回退 CSV"""
    # 尝试 openpyxl（.xlsx）
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_bytes, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, max_col=3, values_only=True))
        wb.close()
        if rows:
            return rows
    except Exception:
        pass

    # 尝试 xlrd（.xls）
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        rows = []
        for i in range(1, ws.nrows):
            rows.append((
                ws.cell_value(i, 0) if ws.ncols > 0 else "",
                ws.cell_value(i, 1) if ws.ncols > 1 else "",
                ws.cell_value(i, 2) if ws.ncols > 2 else "",
            ))
        if rows:
            return rows
    except Exception:
        pass

    # 最终回退：当纯文本解析（CSV / TSV）
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk", errors="ignore")

    first_line = text.split("\n")[0] if text else ""
    if "\t" in first_line:
        delimiter = "\t"
    elif "," in first_line:
        delimiter = ","
    else:
        delimiter = ";"

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    data_rows = all_rows[1:] if len(all_rows) > 1 else all_rows
    return [
        (r[0] if len(r) > 0 else "", r[1] if len(r) > 1 else "", r[2] if len(r) > 2 else "")
        for r in data_rows
    ]


class GyMagnetService:
    """磁力链接管理服务"""

    def import_excel(self, file_bytes: bytes) -> dict[str, int]:
        rows = _read_rows(file_bytes)

        # 收集有效行，文件内部先去重
        seen = set()
        valid: list[dict[str, str]] = []
        for row in rows:
            bangou, title, magnet = row
            if not bangou or not magnet:
                continue
            bangou = str(bangou).strip()
            if bangou in seen:
                continue
            seen.add(bangou)
            valid.append({
                "bangou": bangou,
                "title": str(title or "").strip(),
                "magnet": str(magnet).strip(),
            })

        total = len(valid)
        if not valid:
            return {"imported": 0, "skipped": 0, "total": 0}

        with SessionLocal() as db:
            # 查出数据库中已有的番号
            all_bangou = [v["bangou"] for v in valid]
            existing = set()
            for i in range(0, len(all_bangou), 500):
                batch = all_bangou[i : i + 500]
                rows_q = (
                    db.query(GyMagnet.bangou)
                    .filter(GyMagnet.bangou.in_(batch))
                    .all()
                )
                existing.update(r[0] for r in rows_q)

            # 只插入数据库中不存在的
            to_insert = [v for v in valid if v["bangou"] not in existing]
            skipped = total - len(to_insert)
            imported = 0

            if to_insert:
                now = datetime.now()
                for v in to_insert:
                    v["created_at"] = now
                # MySQL INSERT IGNORE — 一条 SQL 批量写入，冲突自动跳过
                stmt = insert(GyMagnet).prefix_with("IGNORE")
                # 每 500 条一批
                for i in range(0, len(to_insert), 500):
                    batch = to_insert[i : i + 500]
                    db.execute(stmt, batch)
                    imported += len(batch)
                db.commit()

        logger.info("导入完成: imported=%d, skipped=%d, total=%d", imported, skipped, total)
        return {"imported": imported, "skipped": skipped, "total": total}

    def list_magnets(
        self,
        page: int = 1,
        page_size: int = 20,
        keyword: str | None = None,
    ) -> dict[str, Any]:
        with SessionLocal() as db:
            query = db.query(GyMagnet)
            if keyword:
                pattern = f"%{keyword}%"
                query = query.filter(
                    or_(
                        GyMagnet.bangou.ilike(pattern),
                        GyMagnet.title.ilike(pattern),
                    )
                )
            total = query.count()
            items = (
                query.order_by(GyMagnet.id.desc())
                .offset((page - 1) * page_size)
                .limit(page_size)
                .all()
            )
            return {
                "items": [
                    {
                        "id": m.id,
                        "bangou": m.bangou,
                        "title": m.title,
                        "magnet": m.magnet,
                        "downloaded": m.downloaded,
                        "downloaded_at": str(m.downloaded_at) if m.downloaded_at else "",
                        "created_at": str(m.created_at) if m.created_at else "",
                    }
                    for m in items
                ],
                "total": total,
                "page": page,
                "page_size": page_size,
            }

    def get_magnet(self, bangou: str) -> dict[str, Any] | None:
        with SessionLocal() as db:
            m = db.query(GyMagnet).filter(GyMagnet.bangou == bangou).first()
            if not m:
                return None
            return {
                "id": m.id,
                "bangou": m.bangou,
                "title": m.title,
                "magnet": m.magnet,
                "downloaded": m.downloaded,
                "downloaded_at": str(m.downloaded_at) if m.downloaded_at else "",
                "created_at": str(m.created_at) if m.created_at else "",
            }

    def delete_magnets(self, ids: list[int]) -> int:
        with SessionLocal() as db:
            count = db.query(GyMagnet).filter(GyMagnet.id.in_(ids)).delete()
            db.commit()
            return count


_instance: GyMagnetService | None = None


def get_gy_magnet_service() -> GyMagnetService:
    global _instance
    if _instance is None:
        _instance = GyMagnetService()
    return _instance
